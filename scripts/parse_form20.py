#!/usr/bin/env python3
"""
Parse Form-20 result xlsx for AC018 (2021 / 2024 / 2026) into per-booth CSVs.

Each Form-20 is a transposed table: rows = polling stations, columns =
candidates + summary totals. The three years have DIFFERENT layouts, so the
parser auto-detects column roles by scanning the header rows for marker text:

  - candidate columns: any string in the name row that isn't a known header label
  - 'total valid' col: header cell containing "total of valid" / "total no. votes"
  - 'rejected' col:    header cell containing "rejected"
  - 'NOTA' col:         header cell containing "nota"
  - 'grand total' col:  header cell containing "total" (after others claimed)

Then for every data row (first cell is a station number) we emit:
  station_no, candidate_1..N (by name), total_valid, rejected, nota, total

Outputs (one per cycle):
  data/votes_ac018_<year>.csv     per-booth candidate vote counts
  data/candidates_ac018_<year>.json   candidate name + party (+ alliance) per col

Usage: python scripts/parse_form20.py
"""
import csv
import json
import re
import sys
from pathlib import Path

import openpyxl

BASE = Path(__file__).resolve().parent.parent
ANALYSIS = BASE / "analysis"
DATA = BASE / "data"

CYCLES = {
    2026: "AC018-2026.xlsx",
    2024: "AC018-2024.xlsx",
    2021: "AC018-2021.xlsx",
}

# 2026 Form-20 has no party row. Manual party map (official ECI result, confirmed
# by operator). Applied after parsing so re-runs preserve parties. For other ACs
# with the same gap, add a per-AC map here.
PARTY_OVERRIDES_2026 = {
    "M SURESH": "Bahujan Samaj Party",
    "P K SEKARBABU": "Dravida Munnetra Kazhagam",
    "R FOUZAN SHARIFF": "Naam Tamilar Katchi",
    "R MANOHAR": "All India Anna Dravida Munnetra Kazhagam",
    "SINORA P S ASHOK": "Tamilaga Vettri Kazhagam",
    # all other 2026 candidates are Independents (left blank -> resolved below)
}

# Markers that identify NON-candidate header columns (case-insensitive substring).
HEADER_LABELS = [
    "serial no", "sl. no", "sl no", "polling station", "total of valid",
    "total no. votes", "total no of votes", "no. of rejected", "rejected",
    "nota", "grand total", "total", "no. of tendered", "tendered",
]


def is_header_label(s):
    s = re.sub(r"\s+", " ", str(s).lower()).strip()
    return any(lbl in s for lbl in HEADER_LABELS)


def find_layout(rows):
    """Scan the top rows to find: name_row, party_row, station_col, and a map
    of column-index -> role for the summary columns. Returns a dict.

    The ECI Form-20 has a header row ("Serial No / Polling Station No / No of
    Valid Votes Cast in favour of") followed by the candidate-name row, then
    (in 2021/2024) a party row. We locate the HEADER row first, then take the
    candidate names from the row immediately after it."""
    header_row = party_row = None
    station_col = 0
    for i, r in enumerate(rows[:6]):
        nz = [(j, re.sub(r"\s+", " ", str(v)).lower().strip())
              for j, v in enumerate(r) if v is not None]
        joined = " ".join(v for _, v in nz)
        # header row: contains BOTH a "serial/sl no" label AND "polling station".
        # Strip punctuation so "sl. no." / "sl.no" / "serial no" all match.
        stripped = re.sub(r"[.\s]+", " ", joined)
        if (("serial no" in stripped) or ("sl no" in stripped)) \
                and ("polling station" in joined):
            header_row = i
            break

    # candidate names are in header_row + 1 (always the next row in ECI Form-20)
    # EXCEPT 2021 where candidate names are IN the header row itself (and the
    # row after is parties). Pick whichever of (header_row, header_row+1) has
    # MORE candidate-like strings: short names, not party-like words.
    def candidate_score(row):
        """Higher = more candidate-like. Candidate names are short person names;
        party rows contain words like 'Party'/'Kazhagam'/'Independent'."""
        if row is None:
            return -1
        score = 0
        for v in row:
            if v is None or not isinstance(v, str):
                continue
            s = re.sub(r"\s+", " ", v).strip()
            if is_header_label(s) or len(s) > 40 or len(s.split()) > 5:
                continue
            low = s.lower()
            if any(k in low for k in ["party", "kazhagam", "independent",
                                      "congress", "front", "katchi", "maiam"]):
                score -= 2  # party-like, penalize
            else:
                score += 1  # candidate-like
        return score

    if header_row is not None:
        r0 = candidate_score(rows[header_row])
        r1 = candidate_score(rows[header_row + 1]) if header_row + 1 < len(rows) else -1
        name_row = header_row if r0 > r1 else header_row + 1
    else:
        name_row = None

    # party row: the row AFTER names, if it's full of party-ish strings
    if name_row is not None and name_row + 1 < len(rows):
        nz = [(j, re.sub(r"\s+", " ", str(v)).lower().strip())
              for j, v in enumerate(rows[name_row + 1]) if v is not None]
        parties = [v for _, v in nz if not is_header_label(v) and
                   any(k in v for k in ["party", "kazhagam", "independent",
                        "congress", "front", "katchi", "maiam"])]
        if len(parties) >= 3:
            party_row = name_row + 1

    # station col: which HEADER column literally labels the station number
    if header_row is not None:
        for j, v in [(j, re.sub(r"\s+", " ", str(v)).lower().strip())
                     for j, v in enumerate(rows[header_row]) if v is not None]:
            if "polling station" in v:
                station_col = j

    # summary columns: detect from the HEADER row's label cells
    roles = {}
    if header_row is not None:
        for j, v in enumerate(rows[header_row]):
            if v is None:
                continue
            low = re.sub(r"\s+", " ", str(v)).lower()
            if "total of valid" in low or "total no. of votes" in low \
                    or "total no of valid" in low or "total no. votes" in low:
                roles[j] = "total_valid"
            elif low == "rejected" or "no. of rejected" in low or "no. of rejected votes" in low:
                roles[j] = "rejected"
            elif low == "nota":
                roles[j] = "nota"
            elif low == "total" or "grand total" in low:
                roles[j] = "total"

    return {"name_row": name_row, "party_row": party_row,
            "station_col": station_col, "summary_cols": roles}


def parse_cycle(year, fname):
    path = ANALYSIS / fname
    if not path.exists():
        print(f"  skip {year}: {path} missing")
        return None
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    layout = find_layout(rows)
    nr = layout["name_row"]
    if nr is None:
        print(f"  skip {year}: could not locate header row")
        return None

    # candidate columns = non-header string cells in name_row. Filter out ECI
    # section titles ("No of Valid Votes Cast in favour of") which are too long
    # / too wordy to be a person's name.
    cand_cols = []
    for j, v in enumerate(rows[nr]):
        if v is None or not isinstance(v, str):
            continue
        name = re.sub(r"\s+", " ", v).strip()
        if is_header_label(v):
            continue
        # a candidate name is short (<40 chars) and has <=5 words; section
        # titles like "No of Valid Votes Cast in favour of" exceed both.
        if len(name) > 40 or len(name.split()) > 5:
            continue
        cand_cols.append((j, name))

    # candidate metadata (party if available)
    party_by_col = {}
    if layout["party_row"] is not None:
        for j, v in enumerate(rows[layout["party_row"]]):
            if v is not None:
                party_by_col[j] = re.sub(r"\s+", " ", str(v)).strip()

    candidates = [{"col": j, "name": n, "party": party_by_col.get(j, "")} for j, n in cand_cols]

    # 2026 Form-20 has no party row -> apply manual overrides for the named
    # parties; everything else defaults to Independent (resolvable downstream).
    if year == 2026:
        for c in candidates:
            c["party"] = PARTY_OVERRIDES_2026.get(c["name"], "Independent")

    # data rows: first/second cell is an int station number
    station_col = layout["station_col"]
    out = []
    for r in rows:
        # candidate columns are strings; station col may be at 0 or 1
        sno = r[0] if station_col == 0 else (r[1] if len(r) > 1 else r[0])
        try:
            sno = int(sno)
        except (TypeError, ValueError):
            continue
        rec = {"station_no": sno}
        for j, n in cand_cols:
            v = r[j] if j < len(r) else 0
            try:
                rec[n] = int(v or 0)
            except (TypeError, ValueError):
                rec[n] = 0
        for role, col in [(v, k) for k, v in layout["summary_cols"].items()]:
            val = r[col] if col < len(r) else 0
            try:
                rec[role] = int(val or 0)
            except (TypeError, ValueError):
                rec[role] = 0
        out.append(rec)

    out.sort(key=lambda x: x["station_no"])

    # If no explicit total_valid column was detected, compute it as the sum of
    # candidate votes (robust for years like 2024 where the column is unlabeled
    # in the header row and only named in the candidate-name row).
    has_total = "total_valid" in {v for v in layout["summary_cols"].values()}
    cand_names = [c["name"] for c in candidates]
    if not has_total:
        for rec in out:
            rec["total_valid"] = sum(int(rec.get(n, 0) or 0) for n in cand_names)
        print(f"    computed total_valid as candidate-vote sum (no explicit column)")

    # Sanitize: a few Form-20 rows have a blank/garbage total_valid cell even
    # when the column was detected (e.g. AC018-2026 station 184). If the parsed
    # total_valid is 0 but candidate votes sum > 0, recompute from the sum.
    fixed = 0
    for rec in out:
        cand_sum = sum(int(rec.get(n, 0) or 0) for n in cand_names)
        if (not rec.get("total_valid")) and cand_sum > 0:
            rec["total_valid"] = cand_sum
            fixed += 1
    if fixed:
        print(f"    sanitized {fixed} rows with missing total_valid (recomputed from candidate sum)")

    # write per-booth votes CSV
    DATA.mkdir(exist_ok=True)
    vpath = DATA / f"votes_ac018_{year}.csv"
    summary_fields = sorted({v for v in layout["summary_cols"].values()})
    if "total_valid" not in summary_fields:
        summary_fields = ["total_valid"] + summary_fields
    fields = ["station_no"] + cand_names + summary_fields
    with open(vpath, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(out)

    # write candidate metadata
    cpath = DATA / f"candidates_ac018_{year}.json"
    cpath.write_text(json.dumps(candidates, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"  {year}: {len(out)} stations, {len(candidates)} candidates -> {vpath.name} + {cpath.name}")
    print(f"    summary cols detected: {layout['summary_cols']}")
    # sanity: total_valid should exist
    if "total_valid" not in {r for r in layout["summary_cols"].values()}:
        print(f"    WARNING: no total_valid column detected for {year}")
    return {"candidates": candidates, "votes": out, "layout": layout}


def main():
    print("Parsing Form-20 result sheets for AC018...")
    for year, fname in CYCLES.items():
        parse_cycle(year, fname)


if __name__ == "__main__":
    main()
